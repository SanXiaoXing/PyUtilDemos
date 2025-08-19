import hashlib
import json
import os
from pathlib import Path
from openpyxl import load_workbook


HASH_STORE_FILE=str(Path(__file__).parent / 'sheet_hashes.json')
EXCEL_FILE = str(Path(__file__).parent / 'protocol_template.xlsx')


def get_sheet_hashes(file_path, exclude_sheets=None):
    """
    计算 Excel 中所有 sheet（排除指定名称）的 MD5 哈希
    返回字典: {sheet_name: hash_value}
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"{file_path} 不存在")

    if exclude_sheets is None:
        exclude_sheets = ["使用说明"]  # 默认排除

    wb = load_workbook(file_path, data_only=True)
    hashes = {}

    for sheet_name in wb.sheetnames:
        if sheet_name in exclude_sheets:
            continue

        ws = wb[sheet_name]
        # 拼接所有单元格内容
        content = []
        for row in ws.iter_rows(values_only=True):
            content.append(",".join("" if v is None else str(v) for v in row))
        data_str = "\n".join(content)

        # 计算哈希
        md5 = hashlib.md5(data_str.encode("utf-8"))
        hashes[sheet_name] = md5.hexdigest()

    return hashes


def load_previous_hashes(store_file=HASH_STORE_FILE):
    """从 JSON 文件加载上次的哈希"""
    if os.path.exists(store_file):
        with open(store_file, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}  # 没有历史文件则返回空字典


def save_current_hashes(file_path, sheet_hashes, store_file=HASH_STORE_FILE):
    """保存当前哈希到 JSON 文件"""
    all_hashes = load_previous_hashes(store_file)
    all_hashes[file_path] = sheet_hashes
    with open(store_file, "w", encoding="utf-8") as f:
        json.dump(all_hashes, f, indent=4, ensure_ascii=False)


def check_sheets_modified(exclude_sheets=None, store_file=HASH_STORE_FILE):
    """
    自动加载上次哈希并判断修改
    
    参数:
        file_path: Excel 文件路径
        exclude_sheets: 需要排除的工作表列表（默认 ["使用说明"]）
        store_file: 存储哈希的 JSON 文件路径

    返回:
        changed_sheets: 变化的sheet列表
        current_hashes: 当前哈希字典
    """
    file_path=EXCEL_FILE
    previous_all = load_previous_hashes(store_file)
    previous_hashes = previous_all.get(file_path, {})

    current_hashes = get_sheet_hashes(file_path, exclude_sheets)
    changed_sheets = [
        name for name in current_hashes
        if name not in previous_hashes or previous_hashes[name] != current_hashes[name]
    ]

    # 更新保存
    save_current_hashes(file_path, current_hashes, store_file)
    return changed_sheets



