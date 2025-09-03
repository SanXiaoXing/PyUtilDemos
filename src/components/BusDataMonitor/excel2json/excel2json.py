import json
import hashlib
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

EXCEL_FILE = Path(__file__).parent / 'protocol_template.xlsx'
HASH_FILE = Path(__file__).parent / "protocol_hashes.json"
JSON_DIR = Path(__file__).parent.parent / "protocol"
CONF_FILE = Path(__file__).parent.parent / "config/protocol_config.json"


class ProtocolManager:
    def __init__(self, excel_file=None, json_dir=None, hash_file=None):
        self.excel_file = Path(excel_file) if excel_file else EXCEL_FILE
        self.json_dir = Path(json_dir) if json_dir else JSON_DIR
        self.hash_file = Path(hash_file) if hash_file else HASH_FILE
        self.json_dir.mkdir(exist_ok=True)

    # ---------------- 内部工具函数 ----------------
    def _get_sheet_hashes(self, exclude_sheets=None):
        """计算 Excel 中各 sheet 的 MD5 哈希"""
        if not self.excel_file.exists():
            raise FileNotFoundError(f"{self.excel_file} 不存在")

        if exclude_sheets is None:
            exclude_sheets = ["使用说明"]

        wb = load_workbook(self.excel_file, data_only=True)
        hashes = {}

        for sheet_name in wb.sheetnames:
            if sheet_name in exclude_sheets:
                continue

            ws = wb[sheet_name]
            content = []
            for row in ws.iter_rows(values_only=True):
                content.append(",".join("" if v is None else str(v) for v in row))
            data_str = "\n".join(content)

            md5 = hashlib.md5(data_str.encode("utf-8"))
            hashes[sheet_name] = md5.hexdigest()

        return hashes

    def _load_previous_hashes(self):
        if self.hash_file.exists():
            with open(self.hash_file, "r", encoding="utf-8") as f:
                return json.load(f)
        return {}

    def _save_current_hashes(self, sheet_hashes):
        all_hashes = self._load_previous_hashes()
        all_hashes[str(self.excel_file)] = sheet_hashes
        with open(self.hash_file, "w", encoding="utf-8") as f:
            json.dump(all_hashes, f, indent=4, ensure_ascii=False)

    def _check_sheets_modified(self):
        """返回发生变化的 sheet 列表"""
        previous_all = self._load_previous_hashes()
        previous_hashes = previous_all.get(str(self.excel_file), {})
        current_hashes = self._get_sheet_hashes()

        changed_sheets = [
            name for name in current_hashes
            if name not in previous_hashes or previous_hashes[name] != current_hashes[name]
        ]

        self._save_current_hashes(current_hashes)
        return changed_sheets

    def _load_config_info(self):
        """读取 config sheet，返回 {sheet: {length,version,desc}}"""
        df = pd.read_excel(self.excel_file, sheet_name="config")
        cfg = {}
        for _, row in df.iterrows():
            sheet = str(row["sheet"])
            cfg[sheet] = {
                "length": int(row["length"]),
                "version": str(row["version"]),
                "desc": str(row["desc"]),
            }
        return cfg

    def _export_config_json(self, cfg_dict):
        """将 config 信息导出为 json 文件"""
        out_file = CONF_FILE
        out_file.write_text(json.dumps(cfg_dict, indent=4, ensure_ascii=False), encoding="utf-8")
        print(f"协议 config JSON 已生成：{out_file}")

    def _parse_enum_map(self, enum_str):
        """解析 '0:Idle,1:Active' → dict"""
        if pd.isna(enum_str):
            return {}
        mapping = {}
        for item in str(enum_str).split(","):
            kv = item.split(":")
            if len(kv) == 2:
                mapping[kv[0].strip()] = kv[1].strip()
        return mapping

    def _sheet_to_json(self, sheet_name, protocol_info):
        """将单个 sheet 转换为 JSON 文件，增加空值处理，避免 NaN 转换错误"""
        df = pd.read_excel(self.excel_file, sheet_name=sheet_name)

        # --- 标准化列名 ---
        col_map = {
            "name": "Name",
            "byteoffset": "ByteOffset",
            "bitoffset": "BitOffset",
            "bitlength": "BitLength",
            "type": "Type",
            "description": "Description",
            "enummap/value": "EnumMap/Value",
            "scale": "Scale",
            "offset": "Offset"
        }
        df.columns = [col_map.get(str(c).strip().lower(), str(c).strip()) for c in df.columns]

        required_cols = ["Name", "ByteOffset", "BitOffset", "BitLength", "Type"]
        for col in required_cols:
            if col not in df.columns:
                raise ValueError(f"{sheet_name} 缺少必须的列: {col}")

        # 删除完全空的行
        df = df.dropna(how='all')

        fields = []
        for idx, row in df.iterrows():
            name = str(row.get("Name", "")).strip()
            field_type = str(row.get("Type", "")).strip().lower()

            # 跳过没有 Name 或 Type 的行
            if not name or not field_type:
                print(f"警告：{sheet_name} 第 {idx+2} 行缺少 Name 或 Type，已跳过")
                continue

            # 安全转换整数，如果 NaN 则默认 0
            def safe_int(val, default=0):
                try:
                    return int(val) if pd.notna(val) else default
                except (ValueError, TypeError):
                    return default

            byte_offset = safe_int(row.get("ByteOffset"), 0)
            bit_offset = safe_int(row.get("BitOffset"), 0)
            bit_length = safe_int(row.get("BitLength"), 0)

            field = {
                "name": name,
                "byte_offset": byte_offset,
                "bit_offset": bit_offset,
                "bit_length": bit_length,
                "type": field_type,
                "description": str(row.get("Description", "")).strip(),
            }

            if field_type == "enum":
                enum_str = row.get("EnumMap/Value", "")
                field["map"] = self._parse_enum_map(enum_str)
            elif field_type == "fixed":
                scale_val = row.get("Scale", 1.0)
                offset_val = row.get("Offset", 0.0)
                field["scale"] = float(scale_val) if pd.notna(scale_val) else 1.0
                field["offset"] = float(offset_val) if pd.notna(offset_val) else 0.0

            fields.append(field)

        protocol_json = {
            "protocol_name": sheet_name,
            "protocol_length": protocol_info["length"],
            "version": protocol_info["version"],
            "description": protocol_info["desc"],
            "fields": fields
        }

        out_file = self.json_dir / f"{sheet_name}.json"
        out_file.write_text(json.dumps(protocol_json, indent=4, ensure_ascii=False), encoding="utf-8")
        print(f"协议 JSON 已生成：{out_file}")


    def _excel_to_json(self, target_sheets):
        cfg = self._load_config_info()
        self._export_config_json(cfg)  # 每次都更新 config.json

        for sheet_name in target_sheets:
            if sheet_name not in cfg:
                print(f"警告：config 中未找到 {sheet_name} 的配置信息，跳过")
                continue
            self._sheet_to_json(sheet_name, cfg[sheet_name])

    # ---------------- 对外接口 ----------------
    def run(self, force=False):
        """
        主入口：根据参数决定是否直接生成 JSON
        :param force: True = 重新生成所有 sheet 的 JSON
                      False = 仅生成发生变化的 sheet
        """
        cfg = self._load_config_info()
        self._export_config_json(cfg)

        if force:
            print("强制模式：重新生成所有协议 JSON 文件")
            target_sheets = list(cfg.keys())
        else:
            changed_sheets = self._check_sheets_modified()
            if changed_sheets:
                print("检测到以下 sheet 发生变化：", changed_sheets)
                target_sheets = changed_sheets
            else:
                print("没有协议变化，仅更新 config.json")
                target_sheets = []

        self._excel_to_json(target_sheets)


# ---------------- 使用示例 ----------------
if __name__ == "__main__":
    mgr = ProtocolManager(
        excel_file=EXCEL_FILE,
        json_dir=JSON_DIR,
        hash_file=HASH_FILE
    )
    mgr.run(force=True) #强制生成
    #mgr.run(force=False)
